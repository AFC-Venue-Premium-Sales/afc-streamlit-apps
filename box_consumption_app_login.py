import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import re
import io


def preprocess_preorders(preorders_file):
    """Preprocess the Preorders file to clean and standardize it."""
    preorders = pd.read_excel(preorders_file, header=4)
    preorders = preorders.loc[:, ~preorders.columns.str.contains("Unnamed")]
    preorders.columns = preorders.columns.str.strip().str.replace(" ", "_")
    preorders['Location'] = preorders['Location'].ffill()
    preorders = preorders.dropna(how='all')
    preorders = preorders.dropna(subset=['Event', 'Guest_name'])

    preorders['Box_Number'] = preorders['Location'].apply(
        lambda x: re.search(r'\d+', x).group() if pd.notnull(x) and re.search(r'\d+', x) else None
    )
    preorders['Box_Number'] = preorders['Box_Number'].astype(str).str.strip()
    preorders['Total'] = (
        preorders['Total']
        .replace('[\u00a3,]', '', regex=True)
        .replace('', '0')
        .astype(float, errors='ignore')
    )

    def parse_box_numbers(box_entry):
        match = re.findall(r'\d+', box_entry)
        return match if match else [box_entry]

    preorders['Parsed_Box_Numbers'] = preorders['Box_Number'].apply(parse_box_numbers)

    # Multi-event boxes calculation
    multi_event_boxes = preorders.groupby('Box_Number').filter(lambda x: len(x) > 1)

    # Aggregate totals for multi-event boxes
    aggregated = preorders.groupby('Box_Number', as_index=False).agg({
        'Total': 'sum',  # Sum totals for the box
        'Status': lambda x: ', '.join(x.unique()),  # Combine unique statuses
        'Event': lambda x: ', '.join(x.unique())  # Combine unique events
    })

    return preorders, aggregated, multi_event_boxes


def preprocess_box_log(box_log_file):
    """Preprocess the Box Log file to clean and standardize it."""
    box_log = pd.read_excel(box_log_file, sheet_name="ExecutiveBoxesLog", header=2)
    box_log.columns = [
        "Box Number", "Client Name/ Company", "Box Manager",
        "Pre Order Food (INC VAT)", "On Day Order Food (INC VAT)",
        "On Day Order Liquor (INC VAT)", "On Day Order Soft (INC VAT)",
        "On the Day Staff (INC VAT)", "Grand Total (INC VAT)"
    ]
    box_log.columns = box_log.columns.str.strip()
    box_log['Box Number'] = box_log['Box Number'].astype(str).str.strip()

    for col in ["Pre Order Food (INC VAT)", "On Day Order Food (INC VAT)"]:
        box_log[col] = (
            box_log[col]
            .replace('[\u00a3,]', '', regex=True)
            .replace('', '0')
            .astype(float, errors='ignore')
        )
    return box_log


def process_files(box_log_file, preorders_file):
    """Process Box Log and Preorders files and include conditional formatting."""
    box_log = preprocess_box_log(box_log_file)
    preorders, aggregated, multi_event_boxes = preprocess_preorders(preorders_file)

    wb = load_workbook(box_log_file)
    ws = wb["ExecutiveBoxesLog"]

    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="00FF00", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red highlighting

    total_green = 0
    total_yellow = 0
    total_red = 0
    green_details = []
    yellow_details = []
    red_details = []

    matching_boxes = set(box_log['Box Number']).intersection(preorders['Box_Number'])
    total_matching_boxes = len(matching_boxes)

    for index, row in box_log.iterrows():
        box_number = row['Box Number']
        pre_order_value = row['Pre Order Food (INC VAT)']
        on_day_order_value = row['On Day Order Food (INC VAT)']

        matches = preorders[preorders['Parsed_Box_Numbers'].apply(lambda x: box_number in x)]

        if not matches.empty:
            pre_order_total = matches['Total'].sum()
            status = ', '.join(matches['Status'].unique())

            # Apply Green Highlight if values match
            if abs(pre_order_value - pre_order_total) <= 0.01:
                cell = ws[f"D{index + 4}"]
                cell.fill = green_fill
                total_green += 1
                green_details.append(
                    f"Box {box_number} (Row {index + 4}): Pre Order Food (\u00a3{pre_order_value}) matches Preorders Total (\u00a3{pre_order_total}). Status: {status}. Green applied."
                )
            else:
                # Apply Red Highlight if values don't match
                cell = ws[f"D{index + 4}"]
                cell.fill = red_fill
                total_red += 1
                red_details.append(
                    f"Box {box_number} (Row {index + 4}): Pre Order Food (\u00a3{pre_order_value}) does not match Preorders Total (\u00a3{pre_order_total}). Status: {status}. Red applied."
                )

            # Apply Yellow Highlight for Pending status
            if 'Pending' in status:
                cell = ws[f"D{index + 4}"]
                cell.fill = yellow_fill
                total_yellow += 1
                yellow_details.append(
                    f"Box {box_number} (Row {index + 4}): Marked as Yellow due to Pending status. Yellow applied."
                )
        else:
            # Apply Yellow Highlight only if Pre Order Food has a value greater than 0
            if pre_order_value > 0:
                cell = ws[f"D{index + 4}"]
                cell.fill = yellow_fill
                total_yellow += 1
                yellow_details.append(
                    f"Box {box_number} (Row {index + 4}): Not found in Preorders. Highlighted Yellow due to Pre Order Food value (\u00a3{pre_order_value})."
                )

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output, green_details, yellow_details, red_details, total_matching_boxes, matching_boxes, multi_event_boxes


import streamlit as st

def main():
    # Display the Arsenal crest image
    st.image("assets/arsenal_crest_gold.png", width=150)  # Adjust the width as needed
    st.title("Box Log Processing Tool")
    
    # Instructions Expander
    with st.expander("Instructions and Information", expanded=False):
        st.markdown("""
        **How to Use the Tool:**
        1. **Upload Files:** Use the file uploaders below to upload your Box Log and Pre-orders (https://www.tjhub3.com/Rts_Arsenal_Hospitality/Suites/Reports/PreOrders/Index) files in Excel format.
        2. **Process Files:** After uploading both files, click the 'Process Files' button to analyze the data.
        3. **Download Results:** Once processing is complete, a download button will appear to download the processed Box Log.
        4. To see the metrics on the sidebar again, please click **Process Files** once more.

        **How the Tool Works:**
        - The app compares the 'Pre Order Food' values in the Box Log with the totals in the Pre-orders file.
        - It identifies matching boxes and checks for discrepancies between the two files.
        - Based on the comparison, it applies color highlights to the 'Pre Order Food' column in the Box Log to indicate the status.

        **Color Codes Explanation:**
        - **Green:** The 'Pre Order Food' value in Box Log matches the 'Total' in the Pre-orders file if the Status is "Confirmed."
        - **Yellow:** The 'Pre Order Food' value in Box Log matches the 'Total' in the Pre-orders file if the Status is "Pending."
        - **Yellow:** The box is not found in the Pre-orders file, but the 'Pre Order Food' cell in the Box Log has a value.
        - **Red:** The 'Pre Order Food' value does not match the 'Total' in the Preorders file, regardless of status (key for checking Multi-Event Boxes).
          - **On Red filled cells, it is recommended to check the Pre-Order file and manually update if required.**
        """)

    st.sidebar.header("Summary")

    box_log_file = st.file_uploader("Upload Box Log File", type=["xls", "xlsx"])
    preorders_file = st.file_uploader("Upload Preorders File", type=["xls", "xlsx"])

    if box_log_file and preorders_file:
        if st.button("Process Files"):
            output_file, green_details, yellow_details, red_details, total_matching_boxes, matching_boxes, multi_event_boxes = process_files(
                box_log_file, preorders_file
            )
            st.success("Files processed successfully!")

            st.sidebar.subheader("Matching Boxes Summary")
            st.sidebar.write(f"**Total Matching Boxes:** {total_matching_boxes}")
            st.sidebar.write(f"**Matching Boxes:** {', '.join(sorted(matching_boxes))}")

            st.sidebar.subheader("Green Highlights")
            for detail in green_details:
                st.sidebar.write(f"- {detail}")

            st.sidebar.subheader("Yellow Highlights")
            for detail in yellow_details:
                st.sidebar.write(f"- {detail}")

            st.sidebar.subheader("Red Highlights")
            for detail in red_details:
                st.sidebar.write(f"- {detail}")

            st.sidebar.subheader("Multi-Event Boxes")
            if not multi_event_boxes.empty:
                for box, group in multi_event_boxes.groupby('Box_Number'):
                    st.sidebar.write(f"- Box {box} has multiple events with a total of \u00a3{group['Total'].sum():.2f}")

            st.download_button(
                "Download Processed Box Log",
                data=output_file,
                file_name="Processed_Box_Log.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )



def run():
    """Main function to display the Box Consumption Dashboard."""
    st.title("🏟️ Box Consumption Dashboard")
    st.write("Welcome to the Box Log Processing Tool!")

    # Add other UI components as needed
    st.write("Upload your Box Log and Preorders files to begin processing.")


if __name__ == "__main__":
    main()
