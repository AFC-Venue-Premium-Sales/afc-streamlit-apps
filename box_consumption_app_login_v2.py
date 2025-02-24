import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import re
import io

def preprocess_preorders(preorders_file):
    """Preprocess the Preorders file to clean and standardize it."""
    preorders = pd.read_excel(preorders_file, header=4)
    
    # Remove unnamed columns
    preorders = preorders.loc[:, ~preorders.columns.str.contains("Unnamed")]
    
    # Standardize column names
    preorders.columns = preorders.columns.str.strip().str.replace(" ", "_")
    
    # Unmerge the 'Location' column (copy value down)
    preorders['Location'] = preorders['Location'].ffill()
    
    # Drop fully empty rows
    preorders = preorders.dropna(how='all')
    
    # Drop rows missing critical fields
    preorders = preorders.dropna(subset=['Event', 'Order_type'])
    
    # Extract individual Box Numbers from 'Location'
    def extract_box_numbers(location):
        return re.findall(r'\d+', str(location)) if pd.notnull(location) else []

    preorders['Parsed_Box_Numbers'] = preorders['Location'].apply(extract_box_numbers)

    # Expand multiple boxes into separate rows
    preorders = preorders.explode('Parsed_Box_Numbers')
    preorders.rename(columns={'Parsed_Box_Numbers': 'Box_Number'}, inplace=True)

    # Convert Box Number to string
    preorders['Box_Number'] = preorders['Box_Number'].astype(str).str.strip()
    
    # Clean the 'Total' column
    preorders['Total'] = (
        preorders['Total']
        .replace('[\u00a3,]', '', regex=True)
        .replace('', '0')
        .astype(float, errors='ignore')
    )

    # Aggregate data for Multi-Entry Boxes
    aggregated = preorders.groupby('Box_Number', as_index=False).agg({
        'Total': 'sum',  # Sum totals per box
        'Status': lambda x: ', '.join(x.dropna().unique()),  # Combine unique statuses
        'Event': lambda x: ', '.join(x.dropna().unique())    # Combine unique events
    })

    # ✅ Export cleaned Preorders file for validation
    output_path = "/Users/cmunthali/Documents/PYTHON/SALES_REPORTS/PREPROCESS/Cleaned_Preorders.xlsx"
    aggregated.to_excel(output_path, index=False)
    print(f"✅ Cleaned Preorders file saved successfully at: {output_path}")

    return aggregated

def preprocess_box_log(box_log_file):
    """Preprocess the Box Log file."""
    box_log = pd.read_excel(box_log_file, sheet_name="ExecutiveBoxesLog", header=2)
    
    # Define the expected column names (adjust if your file differs)
    box_log.columns = [
        "Box Number", "Client Name/ Company", "Box Manager",
        "Pre Order Food (INC VAT)", "On Day Order Food (INC VAT)",
        "On Day Order Liquor (INC VAT)", "On Day Order Soft (INC VAT)",
        "On the Day Staff (INC VAT)", "Grand Total (INC VAT)"
    ]
    
    box_log.columns = box_log.columns.str.strip()
    box_log['Box Number'] = box_log['Box Number'].astype(str).str.strip()
    
    # Convert currency columns to float
    for col in ["Pre Order Food (INC VAT)", "On Day Order Food (INC VAT)"]:
        box_log[col] = (
            box_log[col]
            .replace('[\u00a3,]', '', regex=True)
            .replace('', '0')
            .astype(float, errors='ignore')
        )
    
    return box_log

def process_files(box_log_file, preorders_file):
    """Process Box Log and Preorders files while keeping formulas intact."""
    box_log = preprocess_box_log(box_log_file)
    preorders = preprocess_preorders(preorders_file)

    wb = load_workbook(box_log_file, data_only=False)  # Load workbook without removing formulas
    ws = wb["ExecutiveBoxesLog"]

    # Define Conditional Formatting Colors
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")    # Red

    total_green = 0
    total_yellow = 0
    total_red = 0
    green_details = []
    yellow_details = []
    red_details = []

    # Identify matching boxes in both files
    matching_boxes = set(box_log['Box Number'].astype(str).str.strip()).intersection(
        preorders['Box_Number'].astype(str).str.strip()
    )

    for index, row in box_log.iterrows():
        box_number = str(row['Box Number']).strip()
        pre_order_value = row['Pre Order Food (INC VAT)'] or 0
        on_day_order_value = row['On Day Order Food (INC VAT)'] or 0

        # Get the event from column B (assuming "Client Name/ Company" holds the event info)
        event_value = row["Client Name/ Company"]

        # Get matching records from Preorders file
        matches = preorders[preorders['Box_Number'] == box_number]
        status = ', '.join(matches['Status'].unique()) if not matches.empty else ""
        preorders_total = matches['Total'].sum() if not matches.empty else 0

        # Worksheet row offset (headers are rows 1-3, data starts at row 4)
        ws_row = index + 4

        # Yellow Logic:
        # If no matching box OR if status is Pending, then:
        if (matches.empty or 'Pending' in status) and pre_order_value > 0:
            # Case 1: No matching record.
            if matches.empty:
                # For MBM events, do not move; just highlight the Pre Order Food cell yellow.
                if event_value.strip().upper() == "MBM":
                    ws[f"D{ws_row}"].fill = yellow_fill
                    yellow_details.append(
                        f"Box {box_number} (Row {ws_row}): MBM event with no matching Preorder; "
                        f"Pre Order Food (£{pre_order_value}) marked yellow."
                    )
                else:
                    # Otherwise, move the value from Pre Order Food to On Day Order Food.
                    new_on_day_order_value = pre_order_value + on_day_order_value
                    ws[f"D{ws_row}"].value = 0
                    ws[f"E{ws_row}"].value = new_on_day_order_value
                    ws[f"E{ws_row}"].fill = yellow_fill
                    yellow_details.append(
                        f"Box {box_number} (Row {ws_row}): No matching Preorder; moved Pre Order Food (£{pre_order_value}) "
                        f"to On Day Order Food."
                    )
            # Case 2: Matching record exists but status is Pending.
            else:
                # If On Day Order Food already matches the Preorders total, simply mark as yellow.
                if abs(on_day_order_value - preorders_total) < 0.01:
                    ws[f"D{ws_row}"].fill = yellow_fill
                    yellow_details.append(
                        f"Box {box_number} (Row {ws_row}): Status Pending and On Day Order Food already matches Preorder total "
                        f"(£{preorders_total}); marked yellow."
                    )
                else:
                    new_on_day_order_value = pre_order_value + on_day_order_value
                    ws[f"D{ws_row}"].value = 0
                    ws[f"E{ws_row}"].value = new_on_day_order_value
                    ws[f"E{ws_row}"].fill = yellow_fill
                    yellow_details.append(
                        f"Box {box_number} (Row {ws_row}): Status Pending; moved Pre Order Food (£{pre_order_value}) to On Day Order Food."
                    )
            total_yellow += 1

        # Completed Logic:
        elif box_number in matching_boxes and 'Completed' in status:
            current_total = pre_order_value + on_day_order_value

            if abs(current_total - preorders_total) < 0.01:
                # Exact match: set Pre Order Food to Preorders total and clear On Day Order Food.
                ws[f"D{ws_row}"].value = preorders_total
                ws[f"E{ws_row}"].value = 0
                ws[f"D{ws_row}"].fill = green_fill

                total_green += 1
                green_details.append(
                    f"Box {box_number} (Row {ws_row}): Pre Order Food matched exactly (£{preorders_total:.2f})."
                )
            elif current_total > preorders_total:
                # More allocated than needed: set Pre Order Food to Preorders total, keep the difference in On Day.
                difference = current_total - preorders_total
                ws[f"D{ws_row}"].value = preorders_total
                ws[f"E{ws_row}"].value = difference

                ws[f"D{ws_row}"].fill = green_fill
                if difference > 0:
                    ws[f"E{ws_row}"].fill = yellow_fill
                    yellow_details.append(
                        f"Box {box_number} (Row {ws_row}): Leftover £{difference:.2f} remains in On Day Order Food."
                    )
                total_green += 1
                green_details.append(
                    f"Box {box_number} (Row {ws_row}): Set Pre Order Food to £{preorders_total:.2f} (Completed)."
                )
            else:
                # Not enough allocated: highlight as red.
                ws[f"D{ws_row}"].fill = red_fill
                ws[f"E{ws_row}"].fill = red_fill
                total_red += 1
                red_details.append(
                    f"Box {box_number} (Row {ws_row}): Completed but only £{current_total:.2f} allocated vs. Preorders total of £{preorders_total:.2f}."
                )

    # Save changes without overwriting formulas
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return (
        output,
        green_details,
        yellow_details,
        red_details,
        len(matching_boxes),
        matching_boxes,
        preorders
    )


def main():
    st.title("Box Log Processing App")
    
    with st.expander("Instructions and Information", expanded=False):
        st.markdown("""
        **How to Use the App:**
        1. **Upload Files:** Use the file uploaders below to upload your Box Log and Preorders files in Excel format.
        2. **Process Files:** After uploading both files, click the 'Process Files' button to analyze the data.
        3. **Download Results:** Once processing is complete, a download button will appear to download the processed Box Log.
        4. To see the metrics on the side bar again, please click **Process Files**.

        **How the App Works:**
        - The app compares the 'Pre Order Food (INC VAT)' values in the Box Log with the 'Total' in the Preorders file.
        - It identifies matching boxes and checks for discrepancies between the two files.
        - Based on the comparison, it applies color highlights to the Box Log to indicate the status.
        """)

    st.sidebar.header("Summary & Explanations")

    box_log_file = st.file_uploader("Upload Box Log File", type=["xls", "xlsx"])
    preorders_file = st.file_uploader("Upload Preorders File", type=["xls", "xlsx"])

    if box_log_file and preorders_file:
        if st.button("Process Files"):
            (
                output_file,
                green_details,
                yellow_details,
                red_details,
                total_matching_boxes,
                matching_boxes,
                multi_event_boxes
            ) = process_files(box_log_file, preorders_file)

            st.success("Files processed successfully!")

            # ---- Reorganized Sidebar ----
            # 1. Matching Boxes Overview
            st.sidebar.subheader("1. Matching Boxes Overview")
            st.sidebar.write(f"**Total Matching Boxes:** {total_matching_boxes}")
            if matching_boxes:
                st.sidebar.write("**Matching Boxes:**")
                for mb in sorted(matching_boxes):
                    st.sidebar.write(f"- Box {mb}")

            # 2. Color Highlights Explanation
            st.sidebar.subheader("2. Color Highlights Explanation")
            st.sidebar.markdown("""
            - **Green:**  
            Indicates that the pre-order has been **confirmed**.  
            The 'Pre Order Food' value exactly matches the total specified in the Preorders file for boxes with a 'Completed' status.  
            If there’s any extra amount, it remains in the 'On Day Order Food' column.

            - **Yellow:**  
            Signals one of two conditions:
            - The pre-order is still **pending** (i.e., the Preorders file shows a 'Pending' status).
            - There is a **surplus** in 'On Day Order Food' after allocating the confirmed pre-order total.

            - **Red:**  
            Flags a **mismatch or error**.  
            This could mean:
            - The total allocated in the Box Log does not meet the Preorders file total for a 'Completed' order.
            """
            )


            # 3. Detailed Box Moves
            st.sidebar.subheader("3. Detailed Box Moves & Highlights")
            if green_details:
                st.sidebar.write("**Green Highlights**")
                for detail in green_details:
                    st.sidebar.write(f"- {detail}")

            if yellow_details:
                st.sidebar.write("**Yellow Highlights**")
                for detail in yellow_details:
                    st.sidebar.write(f"- {detail}")

            if red_details:
                st.sidebar.write("**Red Highlights**")
                for detail in red_details:
                    st.sidebar.write(f"- {detail}")

            # 4. Multi-Event Boxes
            st.sidebar.subheader("4. Multi-Event Boxes")
            if not multi_event_boxes.empty:
                # Group by Box_Number to show sums
                for box, group in multi_event_boxes.groupby('Box_Number'):
                    total_sum = group['Total'].sum()
                    st.sidebar.write(
                        f"- Box {box} has multiple events with a combined total of £{total_sum:.2f}"
                    )
            else:
                st.sidebar.write("No multi-event boxes found.")

            # ---- End Reorganized Sidebar ----

            # Download button
            st.download_button(
                "Download Processed Box Log",
                data=output_file,
                file_name="Processed_Box_Log.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )


if __name__ == "__main__":
    main()
